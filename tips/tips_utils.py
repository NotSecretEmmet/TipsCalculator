from django.contrib.auth.models import User
from django.utils import timezone
from django.utils.timezone import make_aware
from userprofiles.models import Profile

from tips.models import TipsRun, TipsResults

import pandas as pd
import pytz
import isoweek
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def check_input(export_fn, check_isoweek, tips_amount, overtips_amount):
	''' Checks if the inputted workbook is correct.
		Done by checking if the worksheet name, column headers 
		are equal to what is expected. '''
	wb = xlrd.open_workbook(export_fn.name, file_contents=export_fn.read())
	ws = wb.sheet_by_index(0)
	ws_name = wb.sheet_names()[0]
	WORKSHEET_NAME_NL = 'Gepland gewerkt lijst'
	WORKSHEET_NAME_ENG = 'Scheduled worked list'
	if ws_name == WORKSHEET_NAME_NL:
		COL_NAMES = [
		'datum', 'Personeelsnummer', 'Medewerker', 'Afdeling', 'Planning',
		'Eind', 'uren', 'kosten', 'gewerkt', 'Eind', 'uren', 'kosten'
		]
	elif ws_name == WORKSHEET_NAME_ENG:
		COL_NAMES = [
		'date', 'Personeelsnummer', 'Employee', 'Department', 'Planning',
		'End', 'hours', 'costs', 'worked', 'End', 'hours', 'costs'
		]
	else:
		return True, f'Worksheet name "{ws_name}" not recognized. Please check if the correct file is selected.'
	if not check_column_names(ws, COL_NAMES):
		return True, 'Column names unequal to expected values. Please check if the correct file is selected.'
	if check_isoweek:
		lrow = retreive_last_row(ws)
		last_date = xlrd.xldate.xldate_as_datetime(ws.cell_value(rowx=lrow, colx=0), wb.datemode)
		first_date = xlrd.xldate.xldate_as_datetime(ws.cell_value(rowx=1, colx=0), wb.datemode)
		if isoweek.Week.withdate(last_date) != isoweek.Week.withdate(first_date):
			return True, 'First and last dates of inputted sheet are in different weeks. Please make sure the correct dates are selected in the export, or uncheck the check isoweeks checkbox'
	if overtips_amount > tips_amount:
		return True, 'Inputted overtips are greater than the inputted tips amount'
	if tips_amount == 0:
		return True, 'Please enter a non-zero amount of tips'
	return False, 'no error'

def retreive_last_row(ws):
	''' Retreive the highest rown number in the date column.
		Corrects for zero index rownumber.  '''
	count = 1
	for row in range(ws.nrows):
		if ws.cell_value(rowx=row, colx=0) != "":
			count +=1
		else:
			print(row)
			break
	return count -2

def check_column_names(ws, COL_NAMES):
	''' Check if the column names are equal to expected values. '''
	for i in range(0,len(COL_NAMES)):
		if ws.cell_value(rowx=0, colx=i) != COL_NAMES[i]:
			return False
	return True

def calculate_tips(export_fn, tips, overtips, user):
	df = pd.read_excel(export_fn)
	convert_eng_column_headers(df)
	# Drop unncessary columns.
	drop_cols = [4,5,6,7,8,9,11,12]
	df.drop(df.columns[drop_cols], axis=1)
	# Standardize the appartments.
	df['Afdeling'] = df.Afdeling.str.replace(r'(^.*Keuken.*$)', 'Keuken')
	df['Afdeling'] = df.Afdeling.str.replace(r'(^.*Bediening.*$)', 'Bediening')
	# Sum hours by department.
	boh_hours = df.query("Afdeling == 'Keuken'")["uren.1"].sum()
	foh_hours = df.query("Afdeling == 'Bediening'")["uren.1"].sum()
	# Calculate effective and tip rates per hour for each department.
	eff_tips = tips - overtips
	boh_tipr = round((0.5 * eff_tips) / boh_hours,2)
	foh_tipr = round((0.5 * eff_tips) / foh_hours,2)
	# Get needed dates.
	tz = pytz.timezone('Europe/Amsterdam')
	start_dt = make_aware(df.min(axis=0)[0], tz) 
	end_dt = make_aware(df.max(axis=0)[0], tz) 
	run_dt = timezone.now()

	data = {
		'Total_tips' : tips,
		'Overtips' : overtips,
		'Effective_tips' : eff_tips,
		'BOH_tips_hour' : boh_tipr,
		'FOH_tips_hour' : foh_tipr,
		'BOH_Hours' : boh_hours,
		'FOH_Hours' : foh_hours,
		'Total_Hours' : boh_hours + foh_hours,
		'start_dt' : start_dt,
		'end_dt' :  end_dt,
		'run_dt' : run_dt
	}
	tips_list = []
	# Extract unique IDs
	unique_ids = df.Personeelsnummer.unique()
	for idi in unique_ids:
		bhours = df.query(f"Afdeling == 'Keuken' and Personeelsnummer == '{idi}'")["uren.1"].sum()
		btips = round(boh_tipr * bhours,0)
		fhours = df.query(f"Afdeling == 'Bediening' and Personeelsnummer == '{idi}'")["uren.1"].sum()
		ftips = round(foh_tipr * fhours,0)
		thours = bhours + fhours
		ttips = ftips + btips
		tipsi = { 
			'Personeelsnummer' : idi,
			'Naam' : df.loc[df['Personeelsnummer'] == idi, 'Medewerker'].iloc[0],
			'BOH_Hours' :  bhours,
			'BOH_Tips' : btips,
			'FOH_Hours' : fhours,
			'FOH_Tips' : ftips,
			'Total_Hours' : thours,
			'Total_Tips' : ttips
			}
		tips_list.append(tipsi)
	tips_df = pd.DataFrame(tips_list)
	tips_df = tips_df[['Personeelsnummer','Naam','BOH_Hours','FOH_Hours','Total_Hours' ,'BOH_Tips','FOH_Tips','Total_Tips']]
	tips_df = tips_df.sort_values(['BOH_Hours','FOH_Hours'], ascending=False)
	return data, tips_df

def convert_eng_column_headers(df):
	''' Function to change the Dutch headers to English if applicable'''
	if df.columns[0] == 'date' and df.columns[2] == 'Employee':
		df.columns = [
			'datum', 'Personeelsnummer', 'Medewerker', 'Afdeling', 'Planning', 'Eind',
			'uren', 'kosten', 'gewerkt', 'Eind', 'uren.1', 'kosten', 'Opmerking'
			]

def create_database_entries(user, data, df3):
	''' Create DB entries for tips run and tip results '''
	tipsrun_entry = create_tipsrun_entry(user, data)
	create_tipsresults_entries(user, df3, tipsrun_entry)
	return tipsrun_entry

def create_tipsresults_entries(user, df3, tipsrun_entry):
	''' Iterate through each row of the dataframe, and create a tipsresults db entry for each'''
	for index, row in df3.iterrows():
		tips_results_i = TipsResults.objects.create(
			personeelsnummer=row['Personeelsnummer'],
			naam=row['Naam'], 
			boh_hours=row['BOH_Hours'], 
			foh_hours=row['FOH_Hours'], 
			total_hours=row['Total_Hours'],
			boh_tips=row['BOH_Tips'], 
			foh_tips=row['FOH_Tips'], 
			total_tips=row['Total_Tips'], 
			tipsrun=tipsrun_entry, 
			user=user)
		tips_results_i.save()

def create_tipsrun_entry(user, data):
	''' Create the db entry for the tipsrun'''
	tipsrun_entry = TipsRun.objects.create(
		start_dt= data['start_dt'],
		end_dt= data['end_dt'], 
		run_dt= data['run_dt'],
		boh_hours = data['BOH_Hours'], 
		boh_tipr = data['BOH_tips_hour'], 
		foh_hours = data['FOH_Hours'], 
		foh_tipr = data['FOH_tips_hour'],
		tips_amount= data['Total_tips'], 
		overtips_amounts = data['Overtips'], 
		effective_tips = data['Effective_tips'], 
		locatie = user.profile.locatie, 
		week=isoweek.Week.withdate(data['start_dt']), 
		user = user
	)
	tipsrun_entry.save()
	return tipsrun_entry

def process_tips(filename, tips_amount, overtips_amount, user):
	''' Add description here'''
	data, df = calculate_tips(filename, tips_amount, overtips_amount, user)
	tipsrun_entry = create_database_entries(user, data, df)
	return tipsrun_entry

def create_export_headers(ws):
	''' Creates the headers for the export of the tips run in Excel, and formats them, making them bold  '''
	HEADERS = ['#:', 'Name:', 'BOH Hours:', 'FOH Hours:', 'Total Hours:', 'BOH Tips:', 'FOH Tips:', 'Total Tips:']
	for i, head in enumerate(HEADERS):
		ws.cell(row=1, column=i+1).value = head
	bold_font = Font(bold=True)
	header_row = ws[1]
	for cell in header_row:
		cell.font = bold_font

def create_export_totals_legend(ws):
	''' Creates the legend entries for the export of the tips run in Excel. '''
	LEGENDS = ['Total Tips:', 'OverTips:', 'Effective Tips:', 'BOH Hours:', 'FOH Hours:', 'BOH Tip/Hour:', 'FOH Tip/Hour:']
	for i, leg in enumerate(LEGENDS):
		ws.cell(row=i+2, column=11).value = leg

def add_exports_legend_values(ws, tipsrun):
	''' Populates the totals in the legend for the export of the tips run in Excel.  '''
	ws['L2'] = tipsrun.tips_amount
	ws['L2'].number_format = '€ #,##0' 
	ws['L3'] = tipsrun.overtips_amounts
	ws['L3'].number_format = '€ #,##0' 
	ws['L4'] = '=L2 - L3' #| Effective tips
	ws['L5'] = '=SUM(C:C)' #| BOH HOURS
	ws['L6'] = '=SUM(D:D)' #| FOH HOURS
	ws['L7'] = '=ROUND(L4*0.5/L5, 2)' #| BOH tipr
	ws['L7'].number_format = '€ #,##0.00' 
	ws['L8'] = '=ROUND(L4*0.5/L6, 2)' #| BOH tipr
	ws['L8'].number_format = '€ #,##0.00' 	

def add_export_results(ws, tips_results):
	''' Adds the tips results, formulas, and applys (currency) formatting to the export of the tips run in Excel. '''
	num_results = len(tips_results)
	for i in range(0,num_results):
		ws.cell(row=i+2,column=1).value = tips_results[i].personeelsnummer
		ws.cell(row=i+2,column=2).value = tips_results[i].naam
		ws.cell(row=i+2,column=3).value = tips_results[i].boh_hours
		ws.cell(row=i+2,column=4).value = tips_results[i].foh_hours
		ws.cell(row=i+2,column=5).value = tips_results[i].total_hours
		ws.cell(row=i+2,column=6).value = "=ROUND(L7*OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())),0,-3),0)"
		ws.cell(row=i+2,column=6).number_format = '€ #,##0' 
		ws.cell(row=i+2,column=7).value = "=ROUND(L8*OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())),0,-3),0)"
		ws.cell(row=i+2,column=7).number_format = '€ #,##0' 
		ws.cell(row=i+2,column=8).value = "=OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())),0,-1)+OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())),0,-2)"
		ws.cell(row=i+2,column=8).number_format = '€ #,##0' 	

def set_export_column_widths(ws):
	''' Sets the column widths for the export of the tips run in Excel. '''
	for i in range(1,13):
		column_letter = get_column_letter(i)
		column_dimensions = ws.column_dimensions[column_letter]
		if column_letter == 'B':
			column_dimensions.width = 25
		elif column_letter == 'K':
			column_dimensions.width = 15
		else:
			column_dimensions.width = 10

def export_run_to_excel(tipsrun):
	''' Main function for export to Excel. '''
	tips_results = TipsResults.objects.filter(tipsrun=tipsrun).all()
	wb = Workbook()
	ws = wb.active
	ws.title = tipsrun.locatie + tipsrun.week
	create_export_headers(ws)
	create_export_totals_legend(ws)
	add_exports_legend_values(ws, tipsrun)
	add_export_results(ws, tips_results)
	set_export_column_widths(ws)

	return wb